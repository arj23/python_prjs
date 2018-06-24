import pymongo
import datetime
import urllib
from pandas import DataFrame
import pandas as pd
from collections import OrderedDict

username = 'admin'
password = urllib.parse.quote_plus('abc!@#QWE')

db_address = 'mongodb://'+ username +':' + password + '@137.74.100.108/admin?authSource=admin'

def append_df_to_excel(filename, df, sheetname='sheet1', startrow=None,
                       **to_excel_kwargs):
    """
    Append a DataFrame [df] to existing Excel file [filename]
    into [sheet_name] Sheet.
    If [filename] doesn't exist, then this function will create it.

    Parameters:
      filename : File path or existing ExcelWriter
                 (Example: '/path/to/file.xlsx')
      df : dataframe to save to workbook
      sheetname : Name of sheet which will contain DataFrame.
                   (default: 'Sheet1')
      startrow : upper left cell row to dump data frame.
                 Per default (startrow=None) calculate the last row
                 in the existing DF and write to the next row...
      to_excel_kwargs : arguments which will be passed to `DataFrame.to_excel`
                        [can be dictionary]

    Returns: None
    """
    from openpyxl import load_workbook

    # ignore [engine] parameter if it was passed
    if 'engine' in to_excel_kwargs:
        to_excel_kwargs.pop('engine')

    # create a writer for this month and year
    writer = pd.ExcelWriter(filename, engine='openpyxl')

    try:
        # try to open an existing workbook
        writer.book = load_workbook(filename)
        # get the last row in the existing Excel sheet
        # if it was not specified explicitly
        if not startrow and sheetname in writer.book.get_sheet_names():
            startrow = writer.book.get_sheet_by_name(sheetname).max_row

        # copy existing sheets
        writer.sheets = dict(
            (ws.title, ws) for ws in writer.book.worksheets)
    except FileNotFoundError:
        # file does not exist yet, we will create it
        pass

    if not startrow:
        startrow = 0

    # write out the new sheet
    df.to_excel(writer, sheetname, startrow=startrow, **to_excel_kwargs)
    if startrow != 0:
        df.to_excel(writer, sheetname, startrow=startrow, **to_excel_kwargs, header=False)
    else:
        df.to_excel(writer, sheetname, startrow=startrow, **to_excel_kwargs)

    # save the workbook
    writer.save()

def str_datetime_to_obj_datetime(datetime_str):
    date_str = datetime_str.split("T")[0];
    time_str = datetime_str.split("T")[1];
    year = date_str.split("-")[0]
    month = date_str.split("-")[1]
    day = date_str.split("-")[2]
    hour = time_str.split(":")[0]
    minute = time_str.split(":")[1]
    second = time_str.split(":")[2].split("Z")[0]
    date_time = datetime.datetime(int(year), int(month), int(day), int(hour), int(minute), int(second))
    return date_time

def obj_datetime_to_str_datetime(datetime_obj):
    date_str = str(datetime.date.year) + '-' + str(datetime.date.month) + '-' + str(datetime.date.day)
    time_str = str(datetime.time.hour) + ':' + str(datetime.time.minute) + ':' + str(datetime.time.second)
    datetime_str = date_str + 'T' + time_str + 'Z'
    return datetime_str

def connect_to_mongodb(db_url, db_name, db_collection):
    connection = pymongo.MongoClient(db_url)
    return connection[db_name][db_collection]

def main():
    msgs_collection = connect_to_mongodb(db_address, 'stocktwits', 'suggested_msg')

    end_time = datetime.datetime(2017,12,31,23,59,59)
    start_datetime = end_time - datetime.timedelta(minutes=30)
    min_msg_count = 1000000000000
    max_msg_count = 0
    mean_msg_count = 0
    all_msg_count = 0
    loop_count = 0
    date_list = []
    count_list = []
    while datetime.datetime(2017,1,1,0,0,0) < start_datetime :
        result_curser = msgs_collection.find({'created_at': {'$lt': end_time, '$gte': start_datetime}})
        count = result_curser.count();
        if count is 0:
            result = msgs_collection.find({'created_at': {'$lt': start_datetime}})
            if result.count() is 0:
                break;
        all_msg_count += count;
        if min_msg_count > count:
            min_msg_count = count
        if max_msg_count < count:
            max_msg_count = count
        end_time = start_datetime
        start_datetime = start_datetime - datetime.timedelta(minutes=30)
        loop_count += 1;
        date_list.append(start_datetime)
        count_list.append(count)
    mean_msg_count = all_msg_count/loop_count
    print("Mean = {}".format(mean_msg_count))
    print("Min = {}".format(min_msg_count))
    print("Max = {}".format(max_msg_count))
    print("Number of all messages = {}".format(all_msg_count))
    df = DataFrame(data=OrderedDict({'Date Time': date_list, 'count': count_list}))
    df = df.applymap(lambda x: x.encode('unicode_escape').
                     decode('utf-8') if isinstance(x, str) else x)

    append_df_to_excel("2017.xlsx", df)


if __name__ == "__main__": main()
