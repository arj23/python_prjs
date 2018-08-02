import pymongo
import datetime
import urllib
from pandas import DataFrame
import pandas as pd
from collections import OrderedDict
import numpy

username = 'admin'
password = urllib.parse.quote_plus('9X00DfaColorfullGP(d12@l4l')

db_address = 'mongodb://'+ username +':' + password + '@88.99.153.217:23727/admin?authSource=admin'

def append_df_to_excel(filename, df, sheetname='sheet1', startrow=None,
                       **to_excel_kwargs):
    """
    Append a DataFrame [df] to existing Excel file [filename]
    into [sheet_name] Sheet.
    If [filename] doesn't exist, then this function will create it.

    Parameters:
      filename : File path or existing ExcelWriter
                 (Example: '/path/to/file.xlsx')
      df : dataframe to save to workbook
      sheetname : Name of sheet which will contain DataFrame.
                   (default: 'Sheet1')
      startrow : upper left cell row to dump data frame.
                 Per default (startrow=None) calculate the last row
                 in the existing DF and write to the next row...
      to_excel_kwargs : arguments which will be passed to `DataFrame.to_excel`
                        [can be dictionary]

    Returns: None
    """
    from openpyxl import load_workbook

    # ignore [engine] parameter if it was passed
    if 'engine' in to_excel_kwargs:
        to_excel_kwargs.pop('engine')

    # create a writer for this month and year
    writer = pd.ExcelWriter(filename, engine='openpyxl')

    try:
        # try to open an existing workbook
        writer.book = load_workbook(filename)
        # get the last row in the existing Excel sheet
        # if it was not specified explicitly
        if not startrow and sheetname in writer.book.get_sheet_names():
            startrow = writer.book.get_sheet_by_name(sheetname).max_row

        # copy existing sheets
        writer.sheets = dict(
            (ws.title, ws) for ws in writer.book.worksheets)
    except FileNotFoundError:
        # file does not exist yet, we will create it
        pass

    if not startrow:
        startrow = 0

    # write out the new sheet
    df.to_excel(writer, sheetname, startrow=startrow, **to_excel_kwargs)
    if startrow != 0:
        df.to_excel(writer, sheetname, startrow=startrow, **to_excel_kwargs, header=False)
    else:
        df.to_excel(writer, sheetname, startrow=startrow, **to_excel_kwargs)

    # save the workbook
    writer.save()

def str_datetime_to_obj_datetime(datetime_str):
    date_str = datetime_str.split("T")[0];
    time_str = datetime_str.split("T")[1];
    year = date_str.split("-")[0]
    month = date_str.split("-")[1]
    day = date_str.split("-")[2]
    hour = time_str.split(":")[0]
    minute = time_str.split(":")[1]
    second = time_str.split(":")[2].split("Z")[0]
    date_time = datetime.datetime(int(year), int(month), int(day), int(hour), int(minute), int(second))
    return date_time

def obj_datetime_to_str_datetime(datetime_obj):
    date_str = str(datetime_obj.year) + '-' + str(datetime_obj.month) + '-' + str(datetime_obj.day)
    time_str = str(datetime_obj.hour) + ':' + str(datetime_obj.minute) + ':' + str(datetime_obj.second)
    datetime_str = date_str + 'T' + time_str + 'Z'
    return datetime_str

def connect_to_mongodb (db_url, db_name):
    connection = pymongo.MongoClient(db_url)
    return connection[db_name]

def main():
    db_obj = connect_to_mongodb(db_address, 'stocktwits')
    msgs_collection  = db_obj['all_msgs']
    results_collection  = db_obj['results']

    end_time = datetime.datetime(2017,12,31,23,59,59)
    start_datetime = end_time - datetime.timedelta(days=1)
    min_msg_count = 1000000000000
    max_msg_count = 0
    mean_msg_count = 0
    all_msg_count = 0
    loop_count = 0
    date_list = []
    count_list = []
    while datetime.datetime(2013,1,1,0,0,0) < start_datetime :
        result_curser = msgs_collection.find({'created_at': {'$lt': (end_time - datetime.timedelta(hours=9)), '$gte': (start_datetime+datetime.timedelta(hours=14,minutes=30))}})
        count = result_curser.count();
        all_msg_count += count;
        if min_msg_count > count:
            min_msg_count = count
        if max_msg_count < count:
            max_msg_count = count
        end_time = start_datetime
        loop_count += 1;
        date_list.append(start_datetime+datetime.timedelta(hours=14,minutes=30))
        count_list.append(count)
        ###############################
        result_curser = msgs_collection.find({'created_at': {'$lt': (end_time - datetime.timedelta(hours=8,minutes=30)), '$gte': (start_datetime+datetime.timedelta(hours=15))}})
        count = result_curser.count();
        all_msg_count += count;
        if min_msg_count > count:
            min_msg_count = count
        if max_msg_count < count:
            max_msg_count = count
        end_time = start_datetime
        loop_count += 1;
        date_list.append(start_datetime+datetime.timedelta(hours=14,minutes=30))
        count_list.append(count)
        ###############################
        result_curser = msgs_collection.find({'created_at': {'$lt': (end_time - datetime.timedelta(hours=8)), '$gte': (start_datetime+datetime.timedelta(hours=15,minutes=30))}})
        count = result_curser.count();
        all_msg_count += count;
        if min_msg_count > count:
            min_msg_count = count
        if max_msg_count < count:
            max_msg_count = count
        end_time = start_datetime
        loop_count += 1;
        date_list.append(start_datetime+datetime.timedelta(hours=14,minutes=30))
        count_list.append(count)
        ###############################
        result_curser = msgs_collection.find({'created_at': {'$lt': (end_time - datetime.timedelta(hours=7,minutes=30)), '$gte': (start_datetime+datetime.timedelta(hours=16))}})
        count = result_curser.count();
        all_msg_count += count;
        if min_msg_count > count:
            min_msg_count = count
        if max_msg_count < count:
            max_msg_count = count
        end_time = start_datetime
        loop_count += 1;
        date_list.append(start_datetime+datetime.timedelta(hours=14,minutes=30))
        count_list.append(count)
        ###############################
        result_curser = msgs_collection.find({'created_at': {'$lt': (end_time - datetime.timedelta(hours=7)), '$gte': (start_datetime+datetime.timedelta(hours=16,minutes=30))}})
        count = result_curser.count();
        all_msg_count += count;
        if min_msg_count > count:
            min_msg_count = count
        if max_msg_count < count:
            max_msg_count = count
        end_time = start_datetime
        loop_count += 1;
        date_list.append(start_datetime+datetime.timedelta(hours=14,minutes=30))
        count_list.append(count)
        ###############################
        result_curser = msgs_collection.find({'created_at': {'$lt': (end_time - datetime.timedelta(hours=6,minutes=30)), '$gte': (start_datetime+datetime.timedelta(hours=17))}})
        count = result_curser.count();
        all_msg_count += count;
        if min_msg_count > count:
            min_msg_count = count
        if max_msg_count < count:
            max_msg_count = count
        end_time = start_datetime
        loop_count += 1;
        date_list.append(start_datetime+datetime.timedelta(hours=14,minutes=30))
        count_list.append(count)
        ###############################
        result_curser = msgs_collection.find({'created_at': {'$lt': (end_time - datetime.timedelta(hours=6)), '$gte': (start_datetime+datetime.timedelta(hours=17,minutes=30))}})
        count = result_curser.count();
        all_msg_count += count;
        if min_msg_count > count:
            min_msg_count = count
        if max_msg_count < count:
            max_msg_count = count
        end_time = start_datetime
        loop_count += 1;
        date_list.append(start_datetime+datetime.timedelta(hours=14,minutes=30))
        count_list.append(count)
        ###############################
        result_curser = msgs_collection.find({'created_at': {'$lt': (end_time - datetime.timedelta(hours=5,minutes=30)), '$gte': (start_datetime+datetime.timedelta(hours=18))}})
        count = result_curser.count();
        all_msg_count += count;
        if min_msg_count > count:
            min_msg_count = count
        if max_msg_count < count:
            max_msg_count = count
        end_time = start_datetime
        loop_count += 1;
        date_list.append(start_datetime+datetime.timedelta(hours=14,minutes=30))
        count_list.append(count)
        ###############################
        result_curser = msgs_collection.find({'created_at': {'$lt': (end_time - datetime.timedelta(hours=5)), '$gte': (start_datetime+datetime.timedelta(hours=18,minutes=30))}})
        count = result_curser.count();
        all_msg_count += count;
        if min_msg_count > count:
            min_msg_count = count
        if max_msg_count < count:
            max_msg_count = count
        end_time = start_datetime
        loop_count += 1;
        date_list.append(start_datetime+datetime.timedelta(hours=14,minutes=30))
        count_list.append(count)
        ###############################
        result_curser = msgs_collection.find({'created_at': {'$lt': (end_time - datetime.timedelta(hours=4,minutes=30)), '$gte': (start_datetime+datetime.timedelta(hours=19))}})
        count = result_curser.count();
        all_msg_count += count;
        if min_msg_count > count:
            min_msg_count = count
        if max_msg_count < count:
            max_msg_count = count
        end_time = start_datetime
        loop_count += 1;
        date_list.append(start_datetime+datetime.timedelta(hours=14,minutes=30))
        count_list.append(count)
        ###############################
        result_curser = msgs_collection.find({'created_at': {'$lt': (end_time - datetime.timedelta(hours=4)), '$gte': (start_datetime+datetime.timedelta(hours=19,minutes=30))}})
        count = result_curser.count();
        all_msg_count += count;
        if min_msg_count > count:
            min_msg_count = count
        if max_msg_count < count:
            max_msg_count = count
        end_time = start_datetime
        loop_count += 1;
        date_list.append(start_datetime+datetime.timedelta(hours=14,minutes=30))
        count_list.append(count)
        ###############################
        result_curser = msgs_collection.find({'created_at': {'$lt': (end_time - datetime.timedelta(hours=3,minutes=30)), '$gte': (start_datetime+datetime.timedelta(hours=20))}})
        count = result_curser.count();
        all_msg_count += count;
        if min_msg_count > count:
            min_msg_count = count
        if max_msg_count < count:
            max_msg_count = count
        end_time = start_datetime
        loop_count += 1;
        date_list.append(start_datetime+datetime.timedelta(hours=14,minutes=30))
        count_list.append(count)
        ###############################
        result_curser = msgs_collection.find({'created_at': {'$lt': (end_time - datetime.timedelta(hours=3)), '$gte': (start_datetime+datetime.timedelta(hours=20,minutes=30))}})
        count = result_curser.count();
        all_msg_count += count;
        if min_msg_count > count:
            min_msg_count = count
        if max_msg_count < count:
            max_msg_count = count
        end_time = start_datetime
        loop_count += 1;
        date_list.append(start_datetime+datetime.timedelta(hours=14,minutes=30))
        count_list.append(count)
        start_datetime = start_datetime - datetime.timedelta(days=1)
    mean_msg_count = all_msg_count/loop_count
    print("Mean = {}".format(mean_msg_count))
    print("Min = {}".format(min_msg_count))
    print("Max = {}".format(max_msg_count))
    print("Number of all messages = {}".format(all_msg_count))
    results_collection.insert({'30min_trading_daily_Mean': mean_msg_count, '30min_trading_daily_min': min_msg_count, '30min_trading_daily_max': max_msg_count, '30min_trading_all_count': all_msg_count})
    std_dev = numpy.std(count_list)
    results_collection.insert({'30min_trading_std_dev': std_dev})





if __name__ == "__main__": main()
