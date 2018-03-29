import pymongo
import urllib
from pandas import DataFrame
import pandas as pd

from collections import OrderedDict

username = 'admin'
password = urllib.parse.quote_plus('abc!@#QWE')

db_address = 'mongodb://'+ username +':' + password + '@137.74.100.108/admin?authSource=admin'

def connect_to_mongodb (db_url, db_name):
    connection = pymongo.MongoClient(db_url)
    return connection[db_name]

def append_df_to_excel(filename, df, sheetname='sheet1', startrow=None,
                       **to_excel_kwargs):
    """
    Append a DataFrame [df] to existing Excel file [filename]
    into [sheet_name] Sheet.
    If [filename] doesn't exist, then this function will create it.

    Parameters:
      filename : File path or existing ExcelWriter
                 (Example: '/path/to/file.xlsx')
      df : dataframe to save to workbook
      sheetname : Name of sheet which will contain DataFrame.
                   (default: 'Sheet1')
      startrow : upper left cell row to dump data frame.
                 Per default (startrow=None) calculate the last row
                 in the existing DF and write to the next row...
      to_excel_kwargs : arguments which will be passed to `DataFrame.to_excel`
                        [can be dictionary]

    Returns: None
    """
    from openpyxl import load_workbook

    # ignore [engine] parameter if it was passed
    if 'engine' in to_excel_kwargs:
        to_excel_kwargs.pop('engine')

    # create a writer for this month and year
    writer = pd.ExcelWriter(filename, engine='openpyxl')

    try:
        # try to open an existing workbook
        writer.book = load_workbook(filename)
        # get the last row in the existing Excel sheet
        # if it was not specified explicitly
        if not startrow and sheetname in writer.book.get_sheet_names():
            startrow = writer.book.get_sheet_by_name(sheetname).max_row

        # copy existing sheets
        writer.sheets = dict(
            (ws.title, ws) for ws in writer.book.worksheets)
    except FileNotFoundError:
        # file does not exist yet, we will create it
        pass

    if not startrow:
        startrow = 0

    # write out the new sheet
    df.to_excel(writer, sheetname, startrow=startrow, **to_excel_kwargs)
    if startrow != 0:
        df.to_excel(writer, sheetname, startrow=startrow, **to_excel_kwargs, header=False)
    else:
        df.to_excel(writer, sheetname, startrow=startrow, **to_excel_kwargs)

    # save the workbook
    writer.save()

def main():
    db_obj = connect_to_mongodb(db_address, 'stocktwits')
    bullish_collection  = db_obj['bullish_msg']
    bearish_collection = db_obj['bearish_msg']

    result_bullish = bullish_collection.find()
    result_bearish = bearish_collection.find()

    user_list = []
    date_list = []
    body_list = []
    processed_list = []

    for message in result_bullish :
        user_list.append(message['user'])
        date_list.append(message['created_at'])
        body_list.append(message['boddy'])
        processed_list.append(message['processed_body'])

    df = DataFrame(data=OrderedDict({ 'User': user_list, 'Created Date': date_list, 'Body' : body_list,'Process_body': processed_list}))
    df = df.applymap(lambda x: x.encode('unicode_escape').
                                   decode('utf-8') if isinstance(x, str) else x)

    append_df_to_excel("bullish_processed_messages.xlsx",df)

    user_list = []
    date_list = []
    body_list = []
    processed_list = []

    for message in result_bearish :
        user_list.append(message['user'])
        date_list.append(message['created_at'])
        body_list.append(message['boddy'])
        processed_list.append(message['processed_body'])

    df = DataFrame(data=OrderedDict({ 'User': user_list, 'Created Date': date_list, 'Body' : body_list,'Process_body': processed_list}))
    df = df.applymap(lambda x: x.encode('unicode_escape').
                                   decode('utf-8') if isinstance(x, str) else x)

    append_df_to_excel("bearish_processed_messages.xlsx",df)



if __name__ == "__main__": main()
