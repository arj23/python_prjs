import sys
import requests
import json
from pandas import DataFrame
import pandas as pd

from collections import OrderedDict

def append_df_to_excel(filename, df, sheetname='sheet1', startrow=None,
                       **to_excel_kwargs):
    """
    Append a DataFrame [df] to existing Excel file [filename]
    into [sheet_name] Sheet.
    If [filename] doesn't exist, then this function will create it.

    Parameters:
      filename : File path or existing ExcelWriter
                 (Example: '/path/to/file.xlsx')
      df : dataframe to save to workbook
      sheetname : Name of sheet which will contain DataFrame.
                   (default: 'Sheet1')
      startrow : upper left cell row to dump data frame.
                 Per default (startrow=None) calculate the last row
                 in the existing DF and write to the next row...
      to_excel_kwargs : arguments which will be passed to `DataFrame.to_excel`
                        [can be dictionary]

    Returns: None
    """
    from openpyxl import load_workbook

    # ignore [engine] parameter if it was passed
    if 'engine' in to_excel_kwargs:
        to_excel_kwargs.pop('engine')

    # create a writer for this month and year
    writer = pd.ExcelWriter(filename, engine='openpyxl')

    try:
        # try to open an existing workbook
        writer.book = load_workbook(filename)
        # get the last row in the existing Excel sheet
        # if it was not specified explicitly
        if not startrow and sheetname in writer.book.get_sheet_names():
            startrow = writer.book.get_sheet_by_name(sheetname).max_row

        # copy existing sheets
        writer.sheets = dict(
            (ws.title, ws) for ws in writer.book.worksheets)
    except FileNotFoundError:
        # file does not exist yet, we will create it
        pass

    if not startrow:
        startrow = 0

    # write out the new sheet
    df.to_excel(writer, sheetname, startrow=startrow, **to_excel_kwargs)
    if startrow != 0:
        df.to_excel(writer, sheetname, startrow=startrow, **to_excel_kwargs, header=False)
    else:
        df.to_excel(writer, sheetname, startrow=startrow, **to_excel_kwargs)

    # save the workbook
    writer.save()

def main():
    max_count = 20000
    print(max_count)
    start_url = 'https://api.stocktwits.com/api/2/streams/suggested.json?filter=top&max=103187043'
    api_url = 'https://api.stocktwits.com/api/2/streams/suggested.json?filter=top&max={}'
    req = requests.get(start_url)
    data = json.loads(req.text)
    messages = data.get('messages')
    if(messages) :
        messages = data['messages']
    else :
        print('Receive no message from request, parsing finished')
        return
    parsed_count = len(messages)
    id_list = []
    date_list = []
    user_list = []
    sentiment_list = []
    content_list = []
    likes_list = []
    in_reply_to_message_id_list = []
    parent_message_id_list = []

    for message in messages:
            id_list.append(message['id'])
            date = message['created_at']
            date = date.replace('T', ' ')
            date = date.replace('Z', ' UTC')
            date_list.append(date)
            user_list.append(message['user']['username'])
            if(message['entities']['sentiment']) :
                sentiment_list.append(message['entities']['sentiment']['basic'])
            else :
                sentiment_list.append("N/A")
            if(message.get('likes')) :
                likes_list.append(message['likes']['total'])
            else :
                likes_list.append(0)
            if(message.get('conversation')) :
                if(message['conversation']['parent'] == False) :
                    in_reply_to_message_id_list.append(message['conversation']['in_reply_to_message_id'])
                    parent_message_id_list.append(message['conversation']['parent_message_id'])
                else:
                    in_reply_to_message_id_list.append(message['conversation']['in_reply_to_message_id'])
                    parent_message_id_list.append(message['conversation']['parent_message_id'])
            else :
                in_reply_to_message_id_list.append("N/A")
                parent_message_id_list.append("N/A")

            content_list.append(message['body'])
    if(parsed_count< max_count) :
        print(parsed_count)
        idente = data['messages'][-1]['id']
        calc_ident = False
        while(parsed_count < max_count) :
            if(calc_ident==True) :
                idente = data['messages'][-1]['id']
            req_url = api_url.format(idente)
            req = requests.get(req_url)
            print(req_url)
            data = json.loads(req.text)
            if(data) :
                messages = data.get('messages')
                if (messages):
                    print('data = new_data')
                else:
                    print('Receive no message from request, try again...')
                    calc_ident = False
                    continue
            else:
                print('Receive no data from request, try again...')
                calc_ident = False
                continue
            for message in messages:
                id_list.append(message['id'])
                date = message['created_at']
                date = date.replace('T', ' ')
                date = date.replace('Z', ' UTC')
                date_list.append(date)
                user_list.append(message['user']['username'])
                if(message['entities']['sentiment']) :
                    sentiment_list.append(message['entities']['sentiment']['basic'])
                else :
                    sentiment_list.append("N/A")
                if (message.get('likes')):
                    likes_list.append(message['likes']['total'])
                else:
                    likes_list.append(0)
                content_list.append(message['body'])
                if (message.get('conversation')):
                    if (message['conversation']['parent'] == False):
                        in_reply_to_message_id_list.append(message['conversation']['in_reply_to_message_id'])
                        parent_message_id_list.append(message['conversation']['parent_message_id'])
                    else:
                        in_reply_to_message_id_list.append(message['conversation']['in_reply_to_message_id'])
                        parent_message_id_list.append(message['conversation']['parent_message_id'])
                else:
                    in_reply_to_message_id_list.append("N/A")
                    parent_message_id_list.append("N/A")

            parsed_count += len(messages)
            print(parsed_count)
            calc_ident = True
    print("Threshold of messages satisfied")
    print("Saving to Excel...")
    df = DataFrame(data=OrderedDict({'Date Time': date_list, 'User': user_list, 'Sentiment': sentiment_list, 'likes' : likes_list,'Parent_MessageID': parent_message_id_list, 'Reply_to_MessageID' : in_reply_to_message_id_list, 'Content': content_list}),index=id_list)
    df = df.applymap(lambda x: x.encode('unicode_escape').
                                   decode('utf-8') if isinstance(x, str) else x)

    append_df_to_excel("stocktwits.xlsx",df)



if __name__== "__main__" : main()